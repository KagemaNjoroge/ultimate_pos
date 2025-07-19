import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from company.models import Company
from inventory.models import Inventory
from .models import Category, Product
from rest_framework.response import Response
from .serializers import CategorySerializer, ProductSerializer
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework.viewsets import ModelViewSet
from django.views.decorators.http import require_http_methods
from utils.models import Photo


class CategoryApiViewSet(ModelViewSet):

    queryset = Category.objects.all()
    serializer_class = CategorySerializer

    filterset_fields = ("name", "status")


class ProductApiViewSet(ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    filterset_fields = (
        "name",
        "track_inventory",
        "status",
        "supplier",
    )


@login_required()
@require_http_methods(["GET"])
def categories_list_view(request: HttpRequest) -> HttpResponse:
    context = {
        "categories": Category.objects.all(),
    }
    return render(request, "products/categories.html", context=context)


@login_required()
@require_http_methods(["GET"])
def categories_add_view(request: Request) -> Response:

    context = {
        "category_status": Category.status.field.choices,
    }

    return render(request, "products/categories_add.html", context=context)


@login_required()
@require_http_methods(["GET"])
def categories_update_view(request: HttpRequest, category_id: int) -> HttpResponse:

    category = get_object_or_404(Category, id=category_id)
    serializer = CategorySerializer(instance=category)
    context = {
        "category_status": Category.status.field.choices,
        "category": serializer.data,
    }

    return render(request, "products/categories_update.html", context)


@login_required()
@require_http_methods(["GET"])
def categories_detail_view(request: HttpRequest, category_id: int) -> HttpResponse:
    category = get_object_or_404(Category, id=category_id)
    # products in this category
    products = Product.objects.filter(category=category)

    # Calculate statistics
    total_products = products.count()
    active_products_count = products.filter(status="ACTIVE").count()

    # Calculate average price
    if total_products > 0:
        total_price = sum(product.price for product in products)
        average_price = total_price / total_products
    else:
        average_price = 0

    # Get company for currency symbol
    company = Company.objects.first()
    currency_symbol = company.currency_symbol if company else "Ksh"

    context = {
        "category": category,
        "products": products,
        "currency_symbol": currency_symbol,
        "total_products": total_products,
        "active_products_count": active_products_count,
        "average_price": average_price,
    }

    return render(request, "products/category_details.html", context=context)


@login_required()
@require_http_methods(["GET"])
def products_list_view(request: HttpRequest) -> HttpResponse:
    products = Product.objects.all()

    # Calculate stats
    total_products = products.count()
    active_products = products.filter(status="ACTIVE").count()
    categories_count = products.values("category").distinct().count()

    context = {
        "products": products,
        "total_products": total_products,
        "active_products": active_products,
        "categories_count": categories_count,
    }

    company = Company.objects.first()
    if company:
        context["currency_symbol"] = company.currency_symbol
    else:
        context["currency_symbol"] = "Ksh"

    return render(request, "products/products.html", context=context)


@login_required()
@require_http_methods(["GET"])
def products_add_view(request: HttpRequest) -> HttpResponse:

    context = {
        "product_status": Product.status.field.choices,
        "categories": Category.objects.all().filter(status="ACTIVE"),
        "photos": Photo.objects.all(),
    }
    return render(request, "products/products_add.html", context=context)


@login_required()
@require_http_methods(["GET"])
def products_update_view(request, product_id: str):
    product = get_object_or_404(Product, id=product_id)

    context = {
        "product_status": Product.status.field.choices,
        "product": product,
        "photos": Photo.objects.all(),
        "categories": Category.objects.all().filter(status="ACTIVE"),
    }

    return render(request, "products/products_update.html", context)


@login_required()
@require_http_methods(["GET"])
def search_products_ajax_view(request: HttpRequest) -> HttpResponse:
    query = request.GET.get("term")
    if query:
        products = Product.objects.filter(name__icontains=query)
        serializer = ProductSerializer(products, many=True)
        return JsonResponse(data=serializer.data, safe=False)
    else:
        return JsonResponse(data=[])


@login_required()
def upload_excel_view(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":

        categories = Category.objects.all()
        context = {"categories": categories}
        return render(request, "products/upload_csv_excel.html", context=context)
    elif request.method == "POST":

        try:
            excel_file = request.FILES["excel_file"]
            if not excel_file.name.endswith(".xlsx"):
                messages.error(request, "File is not Excel type")
                return redirect("products:upload_excel")
            # if the file is too large, return
            if excel_file.multiple_chunks():
                messages.error(
                    request,
                    "Uploaded file is too big (%.2f MB)."
                    % (excel_file.size / (1000 * 1000),),
                    extra_tags="warning",
                )
                return redirect("products:upload_excel")

            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            for row in worksheet.iter_rows(
                min_row=2, max_row=worksheet.max_row, min_col=1, max_col=6
            ):

                # Coca Cola A nice soft drink 250 1 1

                product = Product()
                product.name = row[0].value
                product.description = row[1].value
                product.price = row[2].value
                if row[3].value == 1:
                    product.track_inventory = True
                else:
                    product.track_inventory = False
                if row[5].value == 1:
                    product.status = "ACTIVE"
                else:
                    product.status = "INACTIVE"
                categ = row[4].value
                try:
                    category = Category.objects.get(id=categ)
                except Exception as e:
                    category = Category.objects.first()
                product.category = category
                product.save()
                # add to if specified
                if product.track_inventory:
                    inv = Inventory.objects.filter(product=product)
                    if inv.exists():
                        inv = inv.first()
                        inv.quantity += 1
                        inv.save()
                    else:
                        inv = Inventory.objects.create(product=product, quantity=1)
                        inv.save()

                return JsonResponse(
                    {
                        "status": "success",
                        "message": "Processing complete. All products have been added to the database.",
                    },
                    safe=True,
                )
        except Exception as e:

            return JsonResponse(
                {
                    "status": "error",
                    "message": "An error occurred while processing the file. The file did not meet the required format or it is corrupted.",
                }
            )


@login_required()
def download_template(request: HttpRequest):
    file_path = "static/upload_products.xltx"
    with open(file_path, "rb") as file:
        response = HttpResponse(file, content_type="application/vnd.ms-excel")
        response["Content-Disposition"] = 'attachment; filename="upload_products.xltx"'
        return response


@login_required()
@require_http_methods(["GET"])
def product_detail_view(request: HttpRequest, product_id: str) -> HttpResponse:
    product = get_object_or_404(Product, id=product_id)

    # Get inventory information
    try:
        inventory = Inventory.objects.get(product=product)
    except Inventory.DoesNotExist:
        inventory = None

    # Get company for currency symbol
    company = Company.objects.first()
    currency_symbol = company.currency_symbol if company else "Ksh"

    context = {
        "product": product,
        "inventory": inventory,
        "currency_symbol": currency_symbol,
    }
    return render(request, "products/product_details.html", context=context)
